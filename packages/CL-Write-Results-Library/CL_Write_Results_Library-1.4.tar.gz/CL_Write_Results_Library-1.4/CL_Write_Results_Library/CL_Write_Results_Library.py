import os
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell
from robot.libraries.BuiltIn import BuiltIn


class CL_Write_Results_Library(object):
    ROBOT_LISTENER_API_VERSION = 2

    def __init__(self, filepath, sheetname, agent_type):
        """
        This Class is to be run as a listener with the following command in the run configuration:
        --listener
        CPL_Write_Results_Library;PycharmProjects\Partners_International\Data\P@I_Traceability_Matrix.xlsx;req_traceability_matrix;local
        OR
        --listener
        CPL_Write_Results_Library;PycharmProjects\Partners_International\Data\P@I_Traceability_Matrix.xlsx;req_traceability_matrix;serveragent

        Passes in parameters path and filename of the Azure DevOps project requirement
        traceability matrix excel file (parameter: filepath), the requirement traceability matrix
        worksheet name (parameter: sheetname), and the agent type (paramerter: agent_type, valid
        values: local or serveragent). Use agent_type = local, if the automation scripts are to be
        run from a local Azure DevOps build agent or from Jenkins.  Use agent_type = serveragent,
        if the automation scripts are to be run from an Azure DevOps build agent created on the
        server
        """
        file = filepath
        if agent_type != 'serveragent':
            username = os.environ['USERNAME']
            print(username)
            file = "c:\\users\\" + username + "\\" + filepath
        self.filename = file
        self.sheetname = sheetname

    def end_test(self, name, attrs):
        """
        When used as listener, this function is triggered when the test ends and does not require inputted
        parameters.
        This functional formats the requirement traceability matrix worksheet excel file and writes the
        following information to the requirement traceability matrix worksheet excel file:
        1. test case status: PASS or FAIL
        2. the error message for failed test cases
        3. the time it took to run each test case, in milliseconds.
        """
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.get_sheet_by_name(self.sheetname)
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 75
        ws.column_dimensions['G'].width = 75
        ws.column_dimensions['H'].width  = 25
        ws.column_dimensions['I'].width = 70
        ws.column_dimensions['J'].width = 125
        ws.column_dimensions['K'].width = 75
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 20
        for cell in ws["1:1"]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True,vertical = 'top')

        rowEndIndex = ws.max_row + 1

        for x in range(2, rowEndIndex):
            if ws.cell(x, 7).value == name:
                ws.cell(x, 8).value = attrs['status']
                ws.cell(x, 9).value = attrs['message']
                ws.cell(x, 12).value = attrs['elapsedtime']
                break  # exit for loop

        wb.save(self.filename)  # Save the workbook
